# encoding :utf-8
import os
import sqlite3
import traceback

import openpyxl

import EnviromentVar as envi
import Notifer as notif


class DataBases(object):
    def __init__(self):
        self.MAXGEN = envi.KBIS_MAXGEN
        self.MINGEN = envi.KBIS_MINGEN

        notif.output('DataBase 1/5 メモリ内のデータベースを確認しています。')

        try:
            os.remove(envi.DATABASE_PLACE)  # for memory leak interrupt
        except:
            notif.output(traceback.format_exc())

        notif.output('DataBase 2/5 データベースに接続します。')

        self.userBook = envi.USERLIST_DATA_DIRECTRY
        self.moneyBook = envi.MANAGEBOOK_PLACE

        self.db = sqlite3.connect(envi.DATABASE_PLACE)
        self.cursor = self.db.cursor()

        notif.output('DataBase 3/5 データベースに接続しました。テーブルの作成、ユーザの追加を行います。')

        # Table Make
        create_table = '''create table users(gen int,realname TEXT,userId TEXT,money int,remarks TEXT,authority TEXT,UNIQUE (realname,userId)) '''
        self.sql = 'insert into users (gen,realname,userId,money,remarks,authority) values (?,?,?,?,?,?)'
        self.cursor.execute(create_table)

        notif.output('DataBase 4/5 ユーザの追加を終了しました。エクセルファイルを読み込みます。')

        for i in range(self.MINGEN, self.MAXGEN):
            try:
                self.cursor.executemany(self.sql, self.CreateUsersFromSheet(i))
            except KeyError:
                notif.output(traceback.format_exc())

        notif.output('DataBase 5/5 読み込みが完了しました。')
        print('ユーザ全体のリストを表示します。')

        for i in self.cursor.execute('''select * from users'''):
            print(i)
        self.db.commit()
        self.cursor.close()
        linelist = self.Search('at', 'all')

    def CreateUsersFromSheet(self, gen):  # SQLに追加できるように手に入れたデータを変換する
        userList = []
        moneybook = openpyxl.load_workbook(self.moneyBook,read_only=True)
        sheet = moneybook[f'{gen}G']

        userbook = openpyxl.load_workbook(self.userBook)
        sheetTB = userbook[envi.USERLIST_DATA_SHEETNAME]

        for user in range(1, envi.MAXUSER_BY_GEN):
            # moneyCreating
            sum = 0
            for debt in range(envi.START_EVENTS, envi.MAX_EVENTS):
                if (sheet.cell(row=(user + 3), column=debt).value):
                    try:
                        sum += int(sheet.cell(row=(user + 3), column=debt).value)
                    except TypeError:  # 多分空白の時に発生するので読み飛ばし
                        pass

            # LINEアカウントとの結びつけ
            exist = False
            authority = None

            for i in range(1, 200):
                if (sheet.cell(row=(user + 3), column=2).value == sheetTB.cell(row=(i + 1), column=1).value):
                    userId = sheetTB.cell(row=(i + 1), column=2).value
                    if (sheetTB.cell(row=(i + 1), column=3).value):
                        authority = sheetTB.cell(row=(i + 1), column=3).value
                    exist = True
            if (not exist):
                i = 0
                notif.output(
                    f'管理簿にいて対応リストにいないUserを発見{sheet.cell(row=(user + 3), column=2).value}')

                while (True):
                    if (sheetTB.cell(row=(i + 1), column=1).value):
                        pass
                    else:
                        sheetTB.cell(row=(i + 1), column=1).value = sheet.cell(row=(user + 3), column=2).value
                        notif.output(
                            f'管理簿にいて対応リストにいないUserを追加しました。{sheet.cell(row=(user + 3), column=2).value}')
                        userId = None
                        authority = None
                        break
                    i = i + 1

            #
            if (sheet.cell(row=(user + 3), column=2).value):
                userList.append((gen, sheet.cell(row=(user + 3), column=2).value, userId, sum,
                                 sheet.cell(row=(user + 3), column=5).value, authority))
            else:
                pass
        moneybook.close()
        userbook.save(self.userBook)
        print(userList)
        return userList

    def renew(self):  # 一回全部消すか・・・
        self.db = sqlite3.connect(envi.DATABASE_PLACE)
        self.cursor = self.db.cursor()
        delete_usersql = '''drop table users'''
        delete_linesql = '''drop table LINEExistsUser'''
        try:
            self.cursor.execute(delete_usersql)
            self.cursor.execute(delete_linesql)
        except:
            traceback.print_exc()
        create_table = '''create table if not exists users(gen int,realname TEXT,userId TEXT,money int,remarks TEXT,authority TEXT,UNIQUE (realname,userId)) '''
        self.cursor.execute(create_table)
        self.sql = 'insert into users (gen,realname,userId,money,remarks,authority) values (?,?,?,?,?,?)'
        createLINEUserTable = '''create table if not exists LINEExistsUser(gen int,realname TEXT,userId TEXT,money int,remarks TEXT,authority TEXT,UNIQUE (realname,userId)) '''
        self.cursor.execute(createLINEUserTable)

        workbook = openpyxl.load_workbook(self.moneyBook,read_only=True)
        for i in range(self.MINGEN, self.MAXGEN):
            try:
                sheet = workbook['{0}G'.format(i)]
                self.cursor.executemany(self.sql, self.CreateUsersFromSheet(i))
            except KeyError:
                break

        print('ユーザ全体のリストを表示します。')

        for i in self.cursor.execute('''select * from users'''):
            print(i)

        select_sql = 'select * from users'
        for row in self.cursor.execute(select_sql):
            print(row)
        workbook.save(self.moneyBook)
        self.db.commit()

        self.cursor.close()

    def Search(self, word1: str, word2: str) -> list:
        self.db = sqlite3.connect(envi.DATABASE_PLACE)
        self.cursor = self.db.cursor()
        createLINEUserTable = '''create table if not exists LINEExistsUser(gen int,realname TEXT,userId TEXT,money int,remarks TEXT,authority TEXT,UNIQUE (realname,userId)) '''
        self.cursor.execute(createLINEUserTable)
        uReturnist = []
        select_sql = '''select * from users where userId is not null'''
        for row in self.cursor.execute(select_sql):
            uReturnist.append(row)
            print(row)
        select_sql = '''insert or ignore into LINEExistsUser(gen,realname,userId,money,remarks,authority) values (?,?,?,?,?,?)'''
        self.cursor.executemany(select_sql, uReturnist)
        select_sql = '''select * from LINEExistsUser'''
        for row in self.cursor.execute(select_sql):
            pass
        returnList = []
        # (word1,word2)
        # (at,本名 or lineID or all)
        # (Lthan,money)
        # (Hthan,money)
        # (equals,money)
        # (get,root)
        notif.output(f'Search command is ({word1} {word2})')
        if (word1 == 'at'):
            if (word2 == 'all'):
                at_all_sql = '''select * from LINEExistsUser'''
                for data in self.cursor.execute(at_all_sql):
                    print(data)
                    returnList.append(data)
                return returnList
            else:  # 本名 or lineID
                print('{0}をLINEIDから検索中...'.format(word2))
                lineID_sql = '''select userId from LINEExistsUser'''
                for data in self.cursor.execute(lineID_sql):
                    # なんかカッコとかついてるので取る
                    ddata0 = str(data).replace('(', '')
                    ddata1 = ddata0.replace(')', '')
                    ddata2 = ddata1.replace('\'', '')
                    ddata_final = ddata2.replace(',', '')
                    # ここでddata_finalがアレ
                    if (ddata_final == word2):
                        print('lineIDが一致しました。 そのユーザを取得します。')
                        userSearch_sql = '''select * from LINEExistsUser where userId='{0}' '''.format(word2)
                        if (not self.cursor.execute(userSearch_sql)):
                            raise ValueError('kasu')
                        reacher = False
                        for i in self.cursor.execute(userSearch_sql):
                            if (reacher):
                                raise AssertionError('二つ以上の要素を持ってしまう致命的なエラー')
                            returnList.append(i)
                            reacher = True
                        return returnList
                notif.output('LINE userIDは一致しませんでした。本名から検索します。')
                #
                notif.output('{0}を本名から検索中...'.format(word2))
                realname_sql = '''select realname from LINEExistsUser'''
                for data in self.cursor.execute(realname_sql):
                    # なんかカッコとかついてるので取る
                    ddata0 = str(data).replace('(', '')
                    ddata1 = ddata0.replace(')', '')
                    ddata2 = ddata1.replace('\'', '')
                    ddata_final = ddata2.replace(',', '')
                    # ここでddata_finalがアレ
                    if (ddata_final == word2):
                        notif.output('本名が一致しました。 そのユーザを取得します。')
                        userSearch_sql = '''select * from LINEExistsUser where realname='{0}' '''.format(word2)
                        if (not self.cursor.execute(userSearch_sql)):
                            raise ValueError('kasu')
                        reacher = False
                        for i in self.cursor.execute(userSearch_sql):
                            if (reacher):
                                raise AssertionError('二つ以上の要素を持ってしまう致命的なエラー')
                            returnList.append(i)
                            reacher = True
                        return returnList
        if (word1 == 'Lthan' or word1 == 'Hthan'):
            if (word1 == 'Lthan'):
                comparison = '<'
            else:
                comparison = '>'
            if (type(word2) is str):
                raise ValueError('SQLインジェクション的な操作は禁止されています。\n引数を確認して下さい。')
            call_sql = f'''select * from LINEExistsUser where money{comparison}{word2}'''
            print(call_sql)
            for i in self.cursor.execute(call_sql):
                returnList.append(i)
            return returnList
        if (word1 == 'get'):
            if (word2 == 'root'):
                select_sql = '''select * from LINEExistsUser where authority=='su' '''
                print(select_sql)
                for i in self.cursor.execute(select_sql):
                    returnList.append(i)
                return returnList

        raise KeyError('値見つからない')

    def RegisterOrChanger(self, Rname: str, NewuserId: str, register=False) -> str:
        self.db = sqlite3.connect(envi.DATABASE_PLACE)
        self.cursor = self.db.cursor()
        if (not register):
            try:
                listy = self.Search('at', Rname)
                for list in listy:
                    Rname = str(list[1])
            except:
                return 'あなたの名前は元々データベースに登録されていません。'
        wb = openpyxl.load_workbook(self.userBook)
        sheet = wb[envi.USERLIST_DATA_SHEETNAME]
        for i in range(2, 200):  # const
            if (sheet.cell(row=i, column=1).value != None):
                print(sheet.cell(row=i, column=1).value)
                if (sheet.cell(row=i, column=1).value == Rname):
                    if (register):
                        if (sheet.cell(row=i, column=2).value == None):
                            sheet.cell(row=i, column=2, value=NewuserId)
                            wb.save(self.userBook)
                            return '登録完了しました！'

                        else:
                            wb.save(self.userBook)
                            return 'すでに登録されています。infoコマンドをお使いください。'
                    else:
                        if (sheet.cell(row=i, column=2).value != None):
                            sheet.cell(row=i, column=2, value=NewuserId)
                            wb.save(self.userBook)
                            return '変更完了しました！'
                        else:
                            wb.save(self.userBook)
                            return 'あなたのデータは登録されていません。registerコマンドを使用してデータベースに登録を行ってください。'
            else:
                break
        wb.save(self.userBook)
        self.db.commit()

        self.cursor.close()
        return 'あなたが誰か判別することが出来ませんでした。管理者に確認することをお勧め致します。'
