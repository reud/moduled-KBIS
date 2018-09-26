import openpyxl
import EnviromentVar as envi


class Budget:
    def __init__(self):
        self.wb = openpyxl.load_workbook(envi.MANAGEBOOK_PLACE, data_only=True, read_only=True)
    def renew(self):
        self.sheet = self.wb[envi.BUDGET_SHEETNAME]
        self.budget_team_design_engineering = self.sheet.cell(row=envi.TEAM_BUDGET_START_ROW,
                                                              column=envi.TEAM_BUDGET_START_COLUMN).value
        self.budget_team_wing = self.sheet.cell(row=envi.TEAM_BUDGET_START_ROW + 1,
                                                column=envi.TEAM_BUDGET_START_COLUMN).value
        self.budget_team_cockpit = self.sheet.cell(row=envi.TEAM_BUDGET_START_ROW + 2,
                                                   column=envi.TEAM_BUDGET_START_COLUMN).value
        self.budget_team_joint = self.sheet.cell(row=envi.TEAM_BUDGET_START_ROW + 3,
                                                 column=envi.TEAM_BUDGET_START_COLUMN).value
        self.budget_team_electrical = self.sheet.cell(row=envi.TEAM_BUDGET_START_ROW + 4,
                                                      column=envi.TEAM_BUDGET_START_COLUMN).value
        self.budget_team_design = self.sheet.cell(row=envi.TEAM_BUDGET_START_ROW + 5,
                                                  column=envi.TEAM_BUDGET_START_COLUMN).value
        self.budget_reserve_fund = self.sheet.cell(row=envi.TEAM_BUDGET_START_ROW + 6,
                                                   column=envi.TEAM_BUDGET_START_COLUMN).value
        self.sheet = self.wb[envi.RECEIPTS_AND_EXPENDITURE_SHEETNAME]
        self.receipts_and_expenditure = self.sheet.cell(row=3,column=4).value

        print(self.budget_team_cockpit)

        if not self.budget_team_cockpit or not self.receipts_and_expenditure:
            print('read failed')
            exit(1)

